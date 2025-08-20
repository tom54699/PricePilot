import os
import tempfile

from openpyxl import load_workbook

from src.python import PriceCalculator
from src.python.price_calculator import export_to_excel


def test_create_quote_and_sums():
    calc = PriceCalculator()
    tasks = [
        {"type": "高級開發", "hours": 2, "complexity": "複雜", "risk": "高風險"},
        {"type": "測試", "hours": 5, "complexity": "簡單", "risk": "低風險"},
        {"type": "不存在的類型", "hours": -3},  # negative hours clamped to 0, default rate applies
    ]
    quote = calc.create_quote("ProjX", "ClientY", tasks, {"維護費用": 200})

    # Compute expected values
    # Item1: base 1800, c=1.8, r=1.5 => adj=4860, subtotal=9720
    # Item2: base 600, c=1.0, r=1.0 => adj=600, subtotal=3000
    # Item3: hours -3 -> 0, default rate "中級開發" 1200 -> subtotal 0
    dev_subtotal = 9720 + 3000
    extras = 200
    pre_tax = dev_subtotal + extras
    tax = round(pre_tax * calc.tax_rate)
    total = pre_tax + tax

    assert quote["開發小計"] == dev_subtotal
    assert quote["額外費用"]["額外費用小計"] == extras
    assert quote["稅前總計"] == pre_tax
    assert quote["營業稅"] == tax
    assert quote["含稅總計"] == total
    assert len(quote["報價項目"]) == 2  # third has 0 hours subtotal but still counted; two non-zero expected


def test_export_to_excel_creates_two_sheets():
    calc = PriceCalculator()
    q = calc.create_quote(
        "ProjA",
        "ClientB",
        [{"type": "中級開發", "hours": 1}],
        {"主機費用": 100},
    )
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "quote_test.xlsx")
        out = export_to_excel(q, path)
        assert os.path.exists(out)

        wb = load_workbook(out)
        assert set(wb.sheetnames) == {"摘要", "詳細項目"}
        ws = wb["摘要"]
        assert ws["A1"].value == "專案名稱"
        assert ws["B1"].value == "ProjA"
        ws2 = wb["詳細項目"]
        assert ws2.max_row >= 2  # header + at least one row
